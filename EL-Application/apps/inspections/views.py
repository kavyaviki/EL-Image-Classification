# inspections/views.py
"""
Views for the Inspections app.
Handles all user-facing pages and API endpoints for image upload, review, and results.
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.http import Http404
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from .models import Inspection, HumanOverride
from .ai_client import AIServiceClient
from django.utils import timezone
from django.db.models import Count, Avg, Q
from datetime import timedelta
# Add these imports at the top of views.py with other imports
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from urllib.parse import quote

import logging
from django.conf import settings

import json
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods

logger = logging.getLogger(__name__)


# ============================================================================
# PAGE VIEWS (Return HTML pages)
# ============================================================================

@login_required
def upload_page(request):
    """
    Render the image upload page.
    
    This view displays the upload interface where users can drag & drop images.
    It also shows the 5 most recent uploads for quick reference.
    
    URL: /inspections/upload/
    Template: inspections/upload.html
    
    Returns:
        HttpResponse: Rendered upload page with recent inspections
    """
    # Get the 5 most recent inspections for the current user
    recent_inspections = Inspection.objects.filter(
        uploaded_by=request.user
    ).order_by('-uploaded_at')[:5]
    
    return render(request, 'inspections/upload.html', {
        'recent_inspections': recent_inspections
    })


@login_required
def inspection_list(request):
    """
    Display a paginated list of all inspections for the current user.
    """
    # Get review threshold from settings
    review_threshold = getattr(settings, 'REVIEW_CONFIDENCE_THRESHOLD', 0.8)
    try:
        review_threshold = float(review_threshold)
    except (ValueError, TypeError):
        review_threshold = 0.8
    
    # Start with all inspections for the current user
    inspections = Inspection.objects.filter(uploaded_by=request.user)
    
    # ============================================================
    # APPLY FILTERS
    # ============================================================
    
    # 1. Search by filename
    search = request.GET.get('search')
    if search:
        inspections = inspections.filter(name__icontains=search)
    
    # 2. Status filter
    status_filter = request.GET.get('status')
    if status_filter:
        if status_filter == 'good':
            inspections = inspections.filter(
                status='completed',
                ai_classification='good',
                ai_confidence__gte=review_threshold
            )
        elif status_filter == 'defect':
            inspections = inspections.filter(
                status='completed',
                ai_classification='defect',
                ai_confidence__gte=review_threshold
            )
        elif status_filter == 'review':
            # Review: completed, confidence < threshold, AND not overridden
            inspections = inspections.filter(
                status='completed',
                ai_confidence__lt=review_threshold,
                human_override=False  # Only show items that still need review
            )
        elif status_filter == 'pending':
            inspections = inspections.filter(status__in=['queued', 'processing'])
        elif status_filter == 'failed':
            inspections = inspections.filter(status='failed')
    
    # 3. Date range filter
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if date_from:
        inspections = inspections.filter(uploaded_at__date__gte=date_from)
    if date_to:
        inspections = inspections.filter(uploaded_at__date__lte=date_to)
    
    # Order by most recent first
    inspections = inspections.order_by('-uploaded_at')
    
    # ============================================================
    # ADD DISPLAY STATUS AND DECISION SOURCE
    # ============================================================
    for inspection in inspections:
        # Determine decision source and status
        if inspection.human_override:
            # If overridden by human, use human's classification
            inspection.display_status = inspection.ai_classification  # This will be 'good' or 'defect'
        elif inspection.status == 'completed':
            # Not overridden, use AI classification with confidence check
            if inspection.ai_confidence is not None and inspection.ai_confidence < review_threshold:
                inspection.display_status = 'review'  # Low confidence = needs review
            else:
                inspection.display_status = inspection.ai_classification if inspection.ai_classification else 'completed'
        elif inspection.status in ['queued', 'processing']:
            inspection.display_status = 'pending'
        elif inspection.status == 'failed':
            inspection.display_status = 'failed'
        else:
            inspection.display_status = inspection.status
    
    # ============================================================
    # PAGINATION
    # ============================================================
    paginator = Paginator(inspections, 10)
    page = request.GET.get('page', 1)
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    context = {
        'page_obj': page_obj,
        'review_threshold': review_threshold,
        'status_choices': Inspection.STATUS_CHOICES,
    }
    return render(request, 'inspections/list.html', context)

@login_required
def inspection_detail(request, pk):
    """
    Show detailed view of a single inspection.
    """
    inspection = get_object_or_404(Inspection, pk=pk, uploaded_by=request.user)
    
    # ============================================================
    # AUTO-FETCH RESULTS IF STILL PENDING
    # ============================================================
    if inspection.status in ['queued', 'processing'] and inspection.job_id:
        ai_client = AIServiceClient()
        results = ai_client.get_results_from_s3(
            str(inspection.job_id),
            inspection.name
        )
        
        if results:
            inspection.status = 'completed'
            inspection.ai_classification = results.get('prediction')
            inspection.ai_confidence = results.get('confidence')
            inspection.ai_processed_at = results.get('processed_at')
            inspection.save()
    
    # ============================================================
    # CHECK FOR HUMAN OVERRIDE
    # ============================================================
    try:
        override = inspection.override
    except HumanOverride.DoesNotExist:
        override = None
    
    # ============================================================
    # CHECK WHERE USER CAME FROM (for back button navigation)
    # ============================================================
    from_review_queue = request.GET.get('from_review_queue', 'false') == 'true'
    
    # ============================================================
    # GET NEXT AND PREVIOUS INSPECTIONS FOR NAVIGATION
    # ============================================================
    next_inspection = None
    previous_inspection = None
    current_index = 0
    total_count = 0
    
    # Get the appropriate list based on where user came from
    if from_review_queue:
        # For review queue: get all pending review items
        review_threshold = getattr(settings, 'REVIEW_CONFIDENCE_THRESHOLD', 0.8)
        try:
            review_threshold = float(review_threshold)
        except (ValueError, TypeError):
            review_threshold = 0.8
        
        all_inspections = Inspection.objects.filter(
            uploaded_by=request.user,
            status='completed',
            ai_confidence__lt=review_threshold,
            human_override=False
        ).order_by('-uploaded_at')
    else:
        # For results dashboard: get all completed inspections
        all_inspections = Inspection.objects.filter(
            uploaded_by=request.user,
            status='completed'
        ).order_by('-uploaded_at')
    
    total_count = all_inspections.count()
    inspection_ids = list(all_inspections.values_list('id', flat=True))
    
    if inspection_ids:
        try:
            current_index = inspection_ids.index(inspection.id) + 1  # 1-based index for display
            
            # Get previous inspection (older)
            if current_index - 2 >= 0:
                previous_inspection = Inspection.objects.get(id=inspection_ids[current_index - 2])
            
            # Get next inspection (newer)
            if current_index < len(inspection_ids):
                next_inspection = Inspection.objects.get(id=inspection_ids[current_index])
                
        except ValueError:
            pass
    
    context = {
        'inspection': inspection,
        'override': override,
        'defects': [],
        'from_review_queue': from_review_queue,
        'next_inspection': next_inspection,
        'previous_inspection': previous_inspection,
        'current_index': current_index,
        'total_count': total_count,
    }
    return render(request, 'inspections/detail.html', context)

@login_required
def review_queue(request):
    """
    Display inspections that need human review.
    
    Review items are defined as:
    - Status = 'completed'
    - AI confidence < REVIEW_CONFIDENCE_THRESHOLD (default 80%)
    - human_override = False (only show pending items)
    
    URL: /inspections/review-queue/
    Template: inspections/review_queue.html
    
    Returns:
        HttpResponse: Rendered review queue page
    """
    # Get review threshold from settings
    review_threshold = getattr(settings, 'REVIEW_CONFIDENCE_THRESHOLD', 0.8)
    try:
        review_threshold = float(review_threshold)
    except (ValueError, TypeError):
        review_threshold = 0.8
    
    # ============================================================
    # QUERY FOR PENDING ITEMS NEEDING REVIEW
    # ============================================================
    # Only completed inspections with confidence below threshold AND not reviewed
    inspections = Inspection.objects.filter(
        uploaded_by=request.user,
        status='completed',                # Must be processed by AI
        ai_confidence__lt=review_threshold, # Low confidence = needs review
        human_override=False                # Only show pending, not reviewed
    )
    
    # ============================================================
    # APPLY FILTERS
    # ============================================================
    
    # Search by filename
    search = request.GET.get('search')
    if search:
        inspections = inspections.filter(name__icontains=search)
    
    # Date range filter
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    if date_from:
        inspections = inspections.filter(uploaded_at__date__gte=date_from)
    if date_to:
        inspections = inspections.filter(uploaded_at__date__lte=date_to)
    
    # Order by most recent first
    inspections = inspections.order_by('-uploaded_at')
    
    # ============================================================
    # CALCULATE SUMMARY CARD COUNTS (removed but kept for potential use)
    # ============================================================
    total_count = inspections.count()
    # Note: Since we're only showing pending, reviewed_count is always 0
    reviewed_count = 0
    pending_count = total_count
    
    # ============================================================
    # ADD DISPLAY STATUS FOR TEMPLATE
    # ============================================================
    for inspection in inspections:
        if inspection.ai_confidence is not None and inspection.ai_confidence < review_threshold:
            inspection.display_status = 'review'
        else:
            inspection.display_status = inspection.ai_classification if inspection.ai_classification else 'completed'
    
    # ============================================================
    # PAGINATION - 10 items per page
    # ============================================================
    paginator = Paginator(inspections, 10)
    page = request.GET.get('page', 1)
    
    try:
        page_obj = paginator.page(page)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)
    
    context = {
        'page_obj': page_obj,
        'total_count': total_count,
        'reviewed_count': reviewed_count,
        'pending_count': pending_count,
        'review_threshold': review_threshold,
    }
    return render(request, 'inspections/review_queue.html', context)


@login_required
def analytics_dashboard(request):
    # ====================== CONFIG ======================
    review_threshold = float(getattr(settings, 'REVIEW_CONFIDENCE_THRESHOLD', 0.8))

    # ====================== FILTER ======================
    date_range = request.GET.get('date_range', '7')
    days = int(date_range)

    today = timezone.now().date()
    start_date = today - timedelta(days=days - 1)

    # ====================== BASE QUERY ======================
    base_qs = Inspection.objects.filter(
        uploaded_by=request.user,
        status='completed',
        uploaded_at__date__gte=start_date
    )

    # ====================== KPI CALCULATIONS ======================
    total_panels = base_qs.count()

    good_panels = base_qs.filter(
        Q(ai_classification='good', ai_confidence__gte=review_threshold) |
        Q(human_override=True, override__new_classification='good')
    ).count()

    defective_panels = base_qs.filter(
        Q(ai_classification='defect', ai_confidence__gte=review_threshold) |
        Q(human_override=True, override__new_classification='defect')
    ).count()

    human_review_required = base_qs.filter(
        ai_confidence__lt=review_threshold,
        human_override=False
    ).count()

    def safe_pct(part, total):
        return round((part / total * 100), 1) if total > 0 else 0

    good_percentage = safe_pct(good_panels, total_panels)
    defective_percentage = safe_pct(defective_panels, total_panels)
    human_review_percentage = safe_pct(human_review_required, total_panels)

    # AI Performance Metrics
    total_completed = total_panels
    auto_classified = base_qs.filter(
        ai_confidence__gte=review_threshold, human_override=False
    ).count()
    human_overridden = base_qs.filter(human_override=True).count()

    auto_classification_rate = safe_pct(auto_classified, total_completed)
    human_override_rate = safe_pct(human_overridden, total_completed)

    # ====================== PREVIOUS PERIOD COMPARISON ======================
    prev_start = start_date - timedelta(days=days)
    prev_end = start_date - timedelta(days=1)

    prev_qs = Inspection.objects.filter(
        uploaded_by=request.user,
        status='completed',
        uploaded_at__date__gte=prev_start,
        uploaded_at__date__lte=prev_end
    )

    prev_total = prev_qs.count()
    prev_good = prev_qs.filter(
        Q(ai_classification='good', ai_confidence__gte=review_threshold) |
        Q(human_override=True, override__new_classification='good')
    ).count()
    prev_defective = prev_qs.filter(
        Q(ai_classification='defect', ai_confidence__gte=review_threshold) |
        Q(human_override=True, override__new_classification='defect')
    ).count()
    prev_human = prev_qs.filter(
        ai_confidence__lt=review_threshold, human_override=False
    ).count()

    def pct_change(curr, prev):
        return round(((curr - prev) / prev * 100), 1) if prev > 0 else 0

    total_change = pct_change(total_panels, prev_total)
    good_change = pct_change(good_panels, prev_good)
    defective_change = pct_change(defective_panels, prev_defective)
    review_change = pct_change(human_review_required, prev_human)

    # ====================== TREND DATA ======================
    trend_qs = base_qs.values('uploaded_at__date').annotate(
        good=Count('id', filter=Q(ai_classification='good', ai_confidence__gte=review_threshold) |
                             Q(human_override=True, override__new_classification='good')),
        defective=Count('id', filter=Q(ai_classification='defect', ai_confidence__gte=review_threshold) |
                                   Q(human_override=True, override__new_classification='defect')),
        human=Count('id', filter=Q(ai_confidence__lt=review_threshold, human_override=False))
    ).order_by('uploaded_at__date')

    trend_dict = {entry['uploaded_at__date']: entry for entry in trend_qs}

    chart_labels = []
    good_values = []
    defective_values = []
    human_values = []
    sparkline_values = []

    for i in range(days):
        date = start_date + timedelta(days=i)
        label = date.strftime('%b %d')
        data = trend_dict.get(date, {'good': 0, 'defective': 0, 'human': 0})

        chart_labels.append(label)
        good_values.append(data.get('good', 0))
        defective_values.append(data.get('defective', 0))
        human_values.append(data.get('human', 0))
        sparkline_values.append(data.get('good', 0) + data.get('defective', 0) + data.get('human', 0))

    # ====================== CONTEXT ======================
    context = {
        'date_range': date_range,
        'date_range_options': [
            {'value': '7', 'label': 'Last 7 Days'},
            {'value': '10', 'label': 'Last 10 Days'},
            {'value': '15', 'label': 'Last 15 Days'},
            {'value': '20', 'label': 'Last 20 Days'},
            {'value': '25', 'label': 'Last 25 Days'},
            {'value': '30', 'label': 'Last 30 Days'},
        ],

        # KPI Cards
        'total_panels': total_panels,
        'total_change': total_change,
        'good_percentage': good_percentage,
        'good_panels': good_panels,
        'good_change': good_change,
        'defective_percentage': defective_percentage,
        'defective_panels': defective_panels,
        'defective_change': defective_change,
        'human_review_percentage': human_review_percentage,
        'human_review_required': human_review_required,
        'review_change': review_change,

        # AI Performance Metrics
        'auto_classification_rate': auto_classification_rate,
        'human_override_rate': human_override_rate,

        # Type Distribution (Radar)
        'distribution_labels': ['Good', 'Defective', 'Human Override'],
        'distribution_values': [good_percentage, defective_percentage, human_review_percentage],

        # Charts
        'chart_labels': chart_labels,
        'good_values': good_values,
        'defective_values': defective_values,
        'human_values': human_values,
        'sparkline_values': sparkline_values,
    }

    return render(request, 'inspections/analytics.html', context)

# ============================================================================
# IMAGE SERVING (For private S3 buckets)
# ============================================================================

@login_required
def serve_image(request, inspection_id):
    """
    Serve an image from private S3 bucket using a presigned URL.
    
    This view is used when the S3 bucket is private (block public access = ON).
    It generates a temporary, authenticated URL that allows the user to view
    the image. The URL expires after 1 hour.
    
    URL: /inspections/image/<uuid:inspection_id>/
    
    Args:
        inspection_id: UUID of the inspection whose image to serve
        
    Returns:
        HttpResponseRedirect: Redirects to the presigned URL
        Http404: If image not found or S3 key missing
        
    How it works:
        1. User visits /inspections/image/123/
        2. This view gets the inspection and its s3_key
        3. Generates a temporary presigned URL valid for 1 hour
        4. Redirects the user to that URL
        5. Browser loads the image from the temporary URL
    """
    # Get the inspection, ensuring it belongs to the current user
    inspection = get_object_or_404(Inspection, id=inspection_id, uploaded_by=request.user)
    
    # Check if we have an S3 key for this image
    if not inspection.s3_key:
        raise Http404("Image not found - no S3 key stored")
    
    # Generate presigned URL using the AI client
    ai_client = AIServiceClient()
    presigned_url = ai_client.get_presigned_image_url(inspection.s3_key, expiration=3600)
    
    if presigned_url:
        # Redirect to the temporary URL
        return redirect(presigned_url)
    
    raise Http404("Could not generate image URL")


# ============================================================================
# API ENDPOINTS (Return JSON for AJAX requests)
# ============================================================================

@login_required
@require_POST
def upload_inspection(request):
    """
    API endpoint to handle image uploads.
    """
    # Validate that files were uploaded
    if not request.FILES.getlist('images'):
        return JsonResponse({'error': 'No images uploaded'}, status=400)
    
    files = request.FILES.getlist('images')
    
    # ========== CRITICAL: Reset file pointers to beginning ==========
    # This ensures we read the complete file content
    for file in files:
        file.seek(0)
    
    ai_client = AIServiceClient()
    
    # Submit to AI service
    result = ai_client.submit_images(files)
    
    if not result or not result.get('job_id'):
        return JsonResponse({'error': 'AI service submission failed'}, status=500)
    
    # Create inspection records for each file
    for file in files:
        # Reset pointer again (extra safety)
        file.seek(0)
        
        # Construct the S3 key
        # s3_key = f"uploads/{result['job_id']}/{file.name}"
        s3_key = f"jobs/{result['job_id']}/input/{file.name}"

        Inspection.objects.create(
            name=file.name,
            uploaded_by=request.user,
            job_id=result['job_id'],
            status='queued',
            s3_key=s3_key
        )
    
    return JsonResponse({
        'success': True,
        'job_id': result['job_id'],
        'file_count': result['file_count'],
        'message': result['message'],
        'status_check_url': result['status_check_url'],
    })


@login_required
def check_job_status(request, job_id):
    """
    API endpoint to check the status of a processing job.
    
    URL: /inspections/api/status/<uuid:job_id>/
    
    Args:
        job_id: The job ID to check
        
    Returns:
        JsonResponse: Status information or error
    """
    ai_client = AIServiceClient()
    status = ai_client.check_status(job_id)
    
    if status:
        return JsonResponse(status)
    
    return JsonResponse({'error': 'Job not found'}, status=404)


@login_required
@require_POST
def human_override(request, pk):
    """
    API endpoint to save a human override for an inspection.
    
    This is called when a user submits the override form on the detail page.
    It creates or updates a HumanOverride record and updates the inspection.
    
    URL: /inspections/api/override/<uuid:pk>/
    Method: POST
    
    Expected POST data:
        classification: 'good' or 'defect'
        reason: Why the override is being made
        notes: Optional additional notes
        
    Returns:
        JsonResponse: Success status and message
    """
    inspection = get_object_or_404(Inspection, pk=pk, uploaded_by=request.user)
    
    # Get data from POST
    classification = request.POST.get('classification')  # 'good' or 'defect'
    reason = request.POST.get('reason')
    notes = request.POST.get('notes', '')
    
    # Validate required fields
    if not classification or not reason:
        return JsonResponse({'success': False, 'error': 'Missing required fields'}, status=400)
    
    # Check if override already exists
    try:
        override = HumanOverride.objects.get(inspection=inspection)
        # Update existing override
        override.original_status = inspection.status
        override.original_classification = inspection.ai_classification
        override.new_status = 'completed'
        override.new_classification = classification
        override.reason = reason
        override.notes = notes
        override.overridden_by = request.user
        override.from_date = timezone.now()
        override.save()
        
    except HumanOverride.DoesNotExist:
        # Create new override
        override = HumanOverride.objects.create(
            inspection=inspection,
            overridden_by=request.user,
            original_status=inspection.status,
            original_classification=inspection.ai_classification,
            new_status='completed',
            new_classification=classification,
            reason=reason,
            notes=notes,
            from_date=timezone.now(),
        )
    
    # Update inspection with human's decision
    inspection.human_override = True
    inspection.ai_classification = classification  # Set to human's decision
    inspection.save()
    
    return JsonResponse({
        'success': True,
        'message': 'Override saved successfully',
        'override_id': str(override.id)
    })


@login_required
@require_POST
def update_inspection_status(request, pk):
    """
    API endpoint to update inspection status from review queue.
    
    This is called when a user selects a new status from the dropdown
    in the review queue table.
    
    URL: /inspections/api/update-status/<uuid:pk>/
    Method: POST
    
    Expected POST data:
        status: New status value
        notes: Optional notes
        
    Returns:
        JsonResponse: Success status and message
    """
    inspection = get_object_or_404(Inspection, pk=pk, uploaded_by=request.user)
    
    new_status = request.POST.get('status')
    
    if new_status:
        inspection.status = new_status
        inspection.save()
        
        return JsonResponse({
            'success': True,
            'message': f'Status updated to {new_status}',
            'new_status': new_status
        })
    
    return JsonResponse({
        'success': False,
        'error': 'No status provided'
    }, status=400)

@login_required
def export_inspections(request):
    """
    API endpoint to export selected or filtered inspections to Excel.
    
    URL: /inspections/api/export/
    Method: GET
    
    Query Parameters:
        ids: Comma-separated list of inspection IDs (UUIDs) to export
        search: Optional search term
        status: Optional status filter
        date_from: Optional start date
        date_to: Optional end date
    
    Returns:
        HttpResponse: Excel file download
    """

    # Add detailed logging at the start
    print(f"=" * 50)
    print(f"EXPORT REQUEST RECEIVED")
    print(f"User: {request.user.email}")
    print(f"GET Parameters: {request.GET}")
    print(f"Time: {timezone.now()}")

    # Get review threshold from settings
    review_threshold = getattr(settings, 'REVIEW_CONFIDENCE_THRESHOLD', 0.8)
    try:
        review_threshold = float(review_threshold)
    except (ValueError, TypeError):
        review_threshold = 0.8
    
    # Get selected IDs from query parameter (UUIDs as strings)
    ids_param = request.GET.get('ids', '')
    print(f"IDs parameter: {ids_param}")
    
    selected_ids = []
    if ids_param:
        # Split by comma and keep as strings (UUIDs)
        selected_ids = [id.strip() for id in ids_param.split(',') if id.strip()]
    
    # Start with all inspections for current user
    inspections = Inspection.objects.filter(uploaded_by=request.user)
    
    # PRIORITY: If specific IDs are selected, export only those
    if selected_ids:
        print(f"Exporting {len(selected_ids)} specific selected inspections: {selected_ids}")
        inspections = inspections.filter(id__in=selected_ids)
    else:
        # Apply filters only if no specific IDs are selected
        print("No specific IDs selected, applying filters...")
        
        # Search filter
        search = request.GET.get('search')
        if search:
            inspections = inspections.filter(name__icontains=search)
            print(f"Applied search filter: {search}")
        
        # Status filter
        status_filter = request.GET.get('status')
        if status_filter:
            if status_filter == 'good':
                inspections = inspections.filter(
                    status='completed',
                    ai_classification='good',
                    ai_confidence__gte=review_threshold
                )
                print(f"Applied status filter: good")
            elif status_filter == 'defect':
                inspections = inspections.filter(
                    status='completed',
                    ai_classification='defect',
                    ai_confidence__gte=review_threshold
                )
                print(f"Applied status filter: defect")
            elif status_filter == 'review':
                inspections = inspections.filter(
                    status='completed',
                    ai_confidence__lt=review_threshold,
                    human_override=False
                )
                print(f"Applied status filter: review")
            elif status_filter == 'pending':
                inspections = inspections.filter(status__in=['queued', 'processing'])
                print(f"Applied status filter: pending")
            elif status_filter == 'failed':
                inspections = inspections.filter(status='failed')
                print(f"Applied status filter: failed")
        
        # Date range filter
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        if date_from:
            inspections = inspections.filter(uploaded_at__date__gte=date_from)
            print(f"Applied date_from filter: {date_from}")
        if date_to:
            inspections = inspections.filter(uploaded_at__date__lte=date_to)
            print(f"Applied date_to filter: {date_to}")
    
    # Order by most recent first
    inspections = inspections.order_by('-uploaded_at')
    
    # Get count for logging
    count = inspections.count()
    print(f"Total inspections to export: {count}")

    # Check if any inspections to export
    if not inspections.exists():
        print(f"No inspections found to export for user {request.user.email}")
        return JsonResponse({'error': 'No inspections found to export'}, status=404)
    
    # Create Excel workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Inspection Results"
    
    # Define column headers
    headers = [
        'S.No',
        'File Name',
        'Date',
        'Status',
        'Decision Source',
        'AI Confidence (%)'
    ]
    
    # Define column widths
    column_widths = [8, 40, 20, 15, 20, 15]
    
    # Style definitions
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='3b82f6', end_color='3b82f6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    cell_alignment_center = Alignment(horizontal='center', vertical='center')
    cell_alignment_left = Alignment(horizontal='left', vertical='center')
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        
        # Set column width
        column_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[column_letter].width = column_widths[col_idx - 1]
    
    # Write data rows
    for row_idx, inspection in enumerate(inspections, 2):
        # Determine display status
        if inspection.human_override:
            display_status = inspection.ai_classification if inspection.ai_classification else 'Completed'
        elif inspection.status == 'completed':
            if inspection.ai_confidence is not None and inspection.ai_confidence < review_threshold:
                display_status = 'Review'
            else:
                display_status = inspection.ai_classification if inspection.ai_classification else 'Completed'
        elif inspection.status in ['queued', 'processing']:
            display_status = 'Pending'
        elif inspection.status == 'failed':
            display_status = 'Failed'
        else:
            display_status = inspection.status
        
        # Determine decision source
        if inspection.human_override:
            decision_source = 'Human Override'
        else:
            decision_source = 'AI Classified'
        
        # Format confidence as percentage
        if inspection.ai_confidence:
            confidence = f"{inspection.ai_confidence * 100:.1f}%"
        else:
            confidence = '-'
        
        # Write row data
        row_data = [
            row_idx - 1,  # S.No
            inspection.name,
            inspection.uploaded_at.strftime('%Y-%m-%d %H:%M') if inspection.uploaded_at else '',
            display_status.capitalize(),
            decision_source,
            confidence
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            
            # Apply alignment based on column
            if col_idx == 2:  # File Name column - left align
                cell.alignment = cell_alignment_left
            else:
                cell.alignment = cell_alignment_center
            
            # Apply color coding for status column
            if col_idx == 4:  # Status column
                if display_status.lower() == 'good':
                    cell.fill = PatternFill(start_color='d1fae5', end_color='d1fae5', fill_type='solid')
                elif display_status.lower() == 'defect' or display_status.lower() == 'defective':
                    cell.fill = PatternFill(start_color='fee2e2', end_color='fee2e2', fill_type='solid')
                elif display_status.lower() == 'review':
                    cell.fill = PatternFill(start_color='fed7aa', end_color='fed7aa', fill_type='solid')
                elif display_status.lower() == 'pending':
                    cell.fill = PatternFill(start_color='e2e3e5', end_color='e2e3e5', fill_type='solid')
            
            # Apply color coding for decision source column
            if col_idx == 5:  # Decision Source column
                if decision_source == 'Human Override':
                    cell.fill = PatternFill(start_color='fed7aa', end_color='fed7aa', fill_type='solid')
                elif decision_source == 'AI Classified':
                    cell.fill = PatternFill(start_color='dbeafe', end_color='dbeafe', fill_type='solid')
    
    # Freeze header row
    worksheet.freeze_panes = 'A2'
    
    # Create HTTP response with Excel file
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Set filename with current date
    filename = f"inspection_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{quote(filename)}"'
    
    # Save workbook to response
    workbook.save(response)
    
    if selected_ids:
        print(f"✅ SUCCESS: User {request.user.email} exported {count} selected inspections to Excel")
    else:
        print(f"✅ SUCCESS: User {request.user.email} exported {count} filtered inspections to Excel")
    print(f"📁 Filename: {filename}")
    print(f"=" * 50)

    return response

# ============================================================================
# CONTEXT PROCESSOR (For sidebar badge count)
# ============================================================================

def review_queue_count(request):
    """
    Context processor to add review queue count to all templates.
    
    This allows the sidebar to display a badge showing how many items
    are pending review.
    
    Add to TEMPLATES context_processors in settings.py:
        'apps.inspections.views.review_queue_count'
        
    Returns:
        dict: {'review_queue_count': count}
    """
    if request.user.is_authenticated:
        # Count items with low confidence (need review)
        review_threshold = getattr(settings, 'REVIEW_CONFIDENCE_THRESHOLD', 0.8)
        count = Inspection.objects.filter(
            uploaded_by=request.user,
            status='completed',
            ai_confidence__lt=review_threshold,
            human_override=False  # Only count those not yet overridden
        ).count()
        return {'review_queue_count': count}
    return {'review_queue_count': 0}

# Retry Button Code
@login_required
@require_http_methods(["POST"])
def retry_analyses(request):
    """
    API endpoint to retry AI analysis for selected inspections.
    Just resets the status so they can be processed again.
    """
    try:
        data = json.loads(request.body)
        inspection_ids = data.get('inspection_ids', [])
        
        if not inspection_ids:
            return JsonResponse({'success': False, 'error': 'No inspection IDs provided'}, status=400)
        
        # Get the inspections
        inspections = Inspection.objects.filter(
            id__in=inspection_ids,
            uploaded_by=request.user
        )
        
        if not inspections.exists():
            return JsonResponse({'success': False, 'error': 'No valid inspections found'}, status=404)
        
        successful = []
        failed = []
        
        for inspection in inspections:
            try:
                # Reset the inspection to queued status
                # Keep existing values to avoid null constraint
                if not inspection.ai_classification:
                    inspection.ai_classification = 'pending'
                if inspection.ai_confidence is None:
                    inspection.ai_confidence = 0.0
                    
                inspection.status = 'queued'
                inspection.human_override = False
                
                # If there's a human override record, delete it
                if hasattr(inspection, 'override'):
                    inspection.override.delete()
                
                inspection.save()
                successful.append(str(inspection.id))
                
            except Exception as e:
                logger.error(f"Failed to retry inspection {inspection.id}: {str(e)}")
                failed.append(str(inspection.id))
        
        message = f"Successfully queued {len(successful)} inspection(s) for re-analysis"
        if failed:
            message += f". Failed: {len(failed)} inspection(s)"
        
        return JsonResponse({
            'success': True,
            'message': message,
            'successful_ids': successful,
            'failed_ids': failed
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': 'Invalid JSON data'}, status=400)
    except Exception as e:
        logger.error(f"Error in retry_analyses: {str(e)}")
        return JsonResponse({'success': False, 'error': str(e)}, status=500)